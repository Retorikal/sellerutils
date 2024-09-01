import logging
import openpyxl as opxl
from database.Card import CardDB as CardDB
import os
import json
import configs
from string import Template
from database.Stocks import StockEntry, StockDB

columns = [
  "Pesan Error",
  "Nama Produk",
  "Deskripsi Produk",
  "Kode Kategori",
  "Berat (Gram)",
  "Minimum Pemesanan",
  "Nomor Etalase",
  "Waktu Proses Preorder",
  "Kondisi",
  "Foto Produk 1",
  "Foto Produk 2",
  "Foto Produk 3",
  "Foto Produk 4",
  "Foto Produk 5",
  "URL Video Produk 1",
  "URL Video Produk 2",
  "URL Video Produk 3",
  "SKU Name",
  "Status",
  "Jumlah Stok",
  "Harga (Rp)",
  "Kurir Pengiriman",
  "Asuransi Pengiriman"
]

# templatepath = "workspace/static/templates/templateDigi.svg"


class TokopediaGenerator():
  def __init__(self, xlspath=None):
    if not xlspath:
      xlspath = configs.template_xls_path()

    if os.path.exists(xlspath):
      self.wb = opxl.load_workbook(xlspath)
    else:
      raise FileNotFoundError

    with (open(configs.product_default_path())) as defaults_file:
      self.defaults = json.load(defaults_file)

    for key in self.defaults:
      if type(self.defaults[key]) is list:
        self.defaults[key] = "\n".join(self.defaults[key])

      if "$" in self.defaults[key]:
        self.defaults[key] = Template(self.defaults[key])

  """
  Write the content of a `Stock` object to an uploadable xls file.
  Creates an iterable based on two database files:
  `stockpath`: CSV file containing how much stock is available for a given card ID
  `pricespath`: Dictionary file containing how much should a given card ID sell for
  """

  def generate_from_new_stock(self, stocks: StockDB, outpath: str):
    self.entry_ws = self.wb["ISI Template Impor Produk"]
    for stock in stocks.parse_stock():
      self.__populate_from_stock_entry(stock)
      logging.info(f"TokopediaGenerator.generate_from_new_stock: added {stock.stock} {
                   stock.stock_id} {stock.name} [{stock.subtitle}]")
    self.wb.save(outpath)

  def update_prices(self, prices: CardDB, outpath: str):
    self.entry_ws = self.wb["Ubah - Informasi Penjualan"]
    column_sku = 11
    column_price = 8
    row = 4
    while True:
      sku = self.entry_ws.cell(row, column_sku)
      price = self.entry_ws.cell(row, column_price)

      if sku.value is None or str(price.value)[-3:] != "000":
        break

      price.value = str(prices.get_price_sku(str(sku.value)) * prices.rate)
      row += 1

    self.wb.save(outpath)

  def __populate_from_stock_entry(self, stock: StockEntry):
    """
    Accepts `stock` and adds an appropriate entry to the Tokopedia uploadable xls.
    @param `stock`: `stock` object:
    """
    row = []
    for column in columns:
      value = self.defaults.get(column, "")

      if type(value) is Template:
        value = value.substitute(stock.get_dict())

      row.append(value)
    self.entry_ws.append(row)
